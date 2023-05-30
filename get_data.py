##
# @file get_data.py
# @brief contain all the function called by the main menu



from ctypes import *
import openpyxl


##
# 
class date(Structure):
    _fields_ = [('year', c_int), ('month', c_int), ('day', c_int)]


class job(Structure):
    _fields_ = [('id', c_int), ('processing_time', c_int),  ('release_date', date), ('due_date', date), ('prtf_value', c_float)
                ]
    


c_file = CDLL("./a.so")

dataArray = (job*100)()
c_file.test.argtypes = [POINTER(job), c_int]
c_file.test.restype = POINTER(c_int)


idName = {}
len = 0
dataframe = None
dataframe_active = None


def readFiles(filename):
    global len, idName, dataframe, dataframe_active
    dataframe = openpyxl.load_workbook(filename)
    dataframe_active = dataframe.active
    for i in range(2, dataframe_active.max_row+1):
        idName[len] = dataframe_active.cell(row=i, column=1).value
        dataArray[i-2].id = len
        dataArray[i -
                  2].processing_time = dataframe_active.cell(row=i, column=4).value
        dataArray[i -
                  2].release_date = HandleDate(str(dataframe_active.cell(row=i, column=2).value))
        dataArray[i -
                  2].due_date = HandleDate(str(dataframe_active.cell(row=i, column=3).value))
        len += 1


# array with the excel datas


# create the object date from the string : 01/01/2000 -> Date(01,01,2000)


def HandleDate(str):
    date_split = str.split(" ")[0].split('/')
    new_date = date()
    new_date.day = int(date_split[0])
    new_date.month = int(date_split[1])
    new_date.year = int(date_split[2])
    return new_date


# dict use to associate the id of a task with the name

# number of row in the excel


def print_datas():
    print("\n")
    print('{:<40} {:<12} {:<12} {:<4}'.format("Task name",
          "Release Date", "End Date", "Processing time"))
    for i in range(2, dataframe_active.max_row+1):

        name = dataframe_active.cell(row=i, column=1).value
        release_date = str(dataframe_active.cell(
            row=i, column=2).value).split(" ")[0]
        due_date = str(dataframe_active.cell(
            row=i, column=3).value).split(" ")[0]
        processing_time = str(dataframe_active.cell(row=i, column=4).value)
        print('{:<40} {:<12} {:<12} {:<4}'.format(
            name, release_date, due_date, processing_time))

# find the id associate with a name of a task


def find_id(seach_str, name_id):
    for key, values in name_id.items():
        if seach_str in values:
            return key
    return None

# get the date enter by the user


def checkInput(message, min, max):
    while True:
        try:
            choice = int(
                input("Enter the" + message+" : "))
            if min <= choice <= max:
                break
            else:
                print("The", message, "must be between", min, "and", max)
        except ValueError:
            print("This is not an integer")
    return choice


def GetDate():
    day = checkInput("day", 1, 31)
    month = checkInput("month", 1, 12)
    year = checkInput("year", 2000, 2100)
    new_date = date()
    new_date.day = int(day)
    new_date.month = int(month)
    new_date.year = int(year)
    return (new_date, '/'.join([str(day), str(month), str(year)]))


def change_datas():
    search_str = input(
        "Enter the name of the Task You Want to modifiate datas : ")
    id = find_id(search_str, idName)
    if id == None:
        print("No matching task with the name ")
        return
    while True:
        try:
            choice = int(input(
                "Enter the Number of the field You Want to modifiate : \n 1) Name \n 2) Release Date \n 3) End Date \n 4) Processing Time \n : "))
            if 1 <= choice <= 4:
                break
            else:
                print("The integer must be between 1 and 4")
        except ValueError:
            print("This is not an integer")

    if choice == 1:
        name = input("Enter the new name of the task : ")
        idName[id] = name
        dataframe_active.cell(row=id+2, column=1).value = name
    if choice == 2:
        print("Enter the release date : ")
        date = GetDate()
        # modifiate the excel
        dataframe_active.cell(row=id+2, column=2).value = date[1]
        # modifiate the array
        dataArray[id-1].release_date = date[0]

    if choice == 3:

        print("Enter the deadline date : ")
        date = GetDate()
        dataframe_active.cell(row=id+2, column=3).value = date[1]
        dataArray[id-1].due_date = date[0]

    if choice == 4:
        while True:
            try:
                processing_time = int(input("Enter the number of days : "))
                if 0 <= choice <= 1000:
                    break
                else:
                    print("The integer must be between 0 and 1000")
            except ValueError:
                print("This is not an integer")

        dataframe_active.cell(row=id+2, column=4).value = processing_time
        dataArray[id-1].processing_time = processing_time


def analyse():
    global len

    print("\nBased on the prtf method, the best order to do your tasks is : ")
    order = c_file.test(dataArray, len)
    for i in range(len):
        print(idName[order[i]])


def Exit(filename):
    print("Exit...")
    dataframe.save(filename)
